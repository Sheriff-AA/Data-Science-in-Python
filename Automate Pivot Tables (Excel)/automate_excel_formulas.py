from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

month = 'February'
workbook = load_workbook('barchart.xlsx')
sheet = workbook['Report']

min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

sheet[f'{get_column_letter(min_column)}{max_row+1}'] = 'Total'

sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=25)
sheet['A2'].font = Font('Arial', bold=True, size=15)

workbook.save(f"report_{month}.xlsx")

